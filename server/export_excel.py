"""
server/export_excel.py — Экспорт результатов тестирования в красиво оформленный Excel (.xlsx).
Включает автоматический расчет оценок (5, 4, 3, 2), цветовую индикацию,
статистику группы и автоподбор ширины столбцов.
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any, Sequence

from PySide6.QtCore import QSettings

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


def calculate_grade_5(percent: float | int, g5: int = 90, g4: int = 70, g3: int = 50) -> tuple[int, str]:
    """
    Рассчитывает 5-балльную оценку по проценту правильных ответов.
    Возвращает кортеж (оценка_число, текстовое_описание).
    """
    if percent >= g5:
        return 5, "Отлично"
    elif percent >= g4:
        return 4, "Хорошо"
    elif percent >= g3:
        return 3, "Удовлетворительно"
    else:
        return 2, "Неудовлетворительно"


def parse_score_percent(score_val: Any) -> float:
    """
    Извлекает процент из строки вида '18/20', '90%' или числа.
    """
    if isinstance(score_val, (int, float)):
        return float(score_val)
    s = str(score_val).strip()
    if not s or s == "—":
        return 0.0
    if "%" in s:
        try:
            return float(s.replace("%", "").strip())
        except ValueError:
            return 0.0
    if "/" in s:
        try:
            parts = s.split("/")
            correct = float(parts[0])
            total = float(parts[1])
            if total > 0:
                return (correct / total) * 100.0
        except (ValueError, IndexError):
            return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def export_results_to_xlsx(
    results: Sequence[dict[str, Any]],
    output_path: str | Path,
    title: str = "Ведомость результатов тестирования",
    group_name: str = "",
    test_title: str = "",
) -> str:
    """
    Экспортирует переданные результаты в стилизованный файл .xlsx.
    """
    if not HAS_OPENPYXL:
        raise RuntimeError("Библиотека openpyxl не установлена. Установите: pip install openpyxl")

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    settings = QSettings("EduTest", "Server")
    g5 = settings.value("grade_5_min", 90, type=int)
    g4 = settings.value("grade_4_min", 70, type=int)
    g3 = settings.value("grade_3_min", 50, type=int)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Результаты"
    ws.views.sheetView[0].showGridLines = True

    # --- Цветовая палитра и стили ---
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")  # Dark Blue
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

    zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    thin_border_side = Side(style="thin", color="CBD5E1")
    cell_border = Border(
        left=thin_border_side,
        right=thin_border_side,
        top=thin_border_side,
        bottom=thin_border_side,
    )

    grade_styles = {
        5: {
            "fill": PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid"),  # Light Green
            "font": Font(name="Calibri", size=11, bold=True, color="166534"),
        },
        4: {
            "fill": PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid"),  # Light Blue
            "font": Font(name="Calibri", size=11, bold=True, color="1E40AF"),
        },
        3: {
            "fill": PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),  # Light Amber
            "font": Font(name="Calibri", size=11, bold=True, color="92400E"),
        },
        2: {
            "fill": PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),  # Light Red
            "font": Font(name="Calibri", size=11, bold=True, color="991B1B"),
        },
    }

    # --- 1. Шапка документа (Title Block) ---
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = title
    title_cell.font = Font(name="Calibri", size=15, bold=True, color="0F172A")
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:H2")
    sub_cell = ws["A2"]
    now_str = datetime.now().strftime("%d.%m.%Y %H:%M")
    sub_info = f"Дата формирования: {now_str}"
    if group_name:
        sub_info += f"   |   Группа: {group_name}"
    if test_title:
        sub_info += f"   |   Тест: {test_title}"
    sub_cell.value = sub_info
    sub_cell.font = Font(name="Calibri", size=10, italic=True, color="64748B")
    sub_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 20

    ws.append([])  # Строка 3
    ws.row_dimensions[3].height = 10

    # --- 2. Заголовки таблицы (Строка 4) ---
    headers = [
        "№",
        "ФИО студента",
        "Группа",
        "Тест / Дисциплина",
        "Баллы",
        "Процент",
        "Оценка",
        "Дата и время",
    ]
    ws.append(headers)
    header_row_idx = 4
    ws.row_dimensions[header_row_idx].height = 26

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row_idx, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = cell_border

    # --- 3. Строки с результатами студентов ---
    row_idx = 5
    grade_counts = {5: 0, 4: 0, 3: 0, 2: 0}
    total_percents = []

    for idx, item in enumerate(results, start=1):
        name = str(item.get("name", "")).strip()
        grp = str(item.get("group", "")).strip()
        t_name = str(item.get("test_name", "")).strip()
        score_str = str(item.get("score", "")).strip()
        ts = str(item.get("timestamp", "")).strip()

        pct = parse_score_percent(score_str)
        total_percents.append(pct)
        grade_num, _grade_desc = calculate_grade_5(pct, g5=g5, g4=g4, g3=g3)
        grade_counts[grade_num] += 1

        row_fill = zebra_fill if idx % 2 == 0 else white_fill

        row_values = [
            idx,
            name,
            grp,
            t_name or "—",
            score_str or "0/0",
            f"{int(pct)}%",
            grade_num,
            ts or "—",
        ]
        ws.append(row_values)
        ws.row_dimensions[row_idx].height = 22

        for col_idx in range(1, len(headers) + 1):
            c = ws.cell(row=row_idx, column=col_idx)
            c.border = cell_border
            c.font = Font(name="Calibri", size=11, color="1E293B")
            c.fill = row_fill

            if col_idx == 1:  # №
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.font = Font(name="Calibri", size=10, color="64748B")
            elif col_idx == 2:  # ФИО
                c.alignment = Alignment(horizontal="left", vertical="center")
                c.font = Font(name="Calibri", size=11, bold=True, color="0F172A")
            elif col_idx in (3, 4, 5, 8):  # Группа, Тест, Баллы, Дата
                c.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx == 6:  # Процент
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.font = Font(name="Calibri", size=11, bold=True, color="0F172A")
            elif col_idx == 7:  # Оценка
                c.alignment = Alignment(horizontal="center", vertical="center")
                g_style = grade_styles.get(grade_num, grade_styles[2])
                c.fill = g_style["fill"]
                c.font = g_style["font"]

        row_idx += 1

    # --- 4. Блок итоговой статистики (Summary Stats) ---
    ws.append([])
    ws.row_dimensions[row_idx].height = 14
    row_idx += 1

    total_count = len(results)
    avg_pct = (sum(total_percents) / total_count) if total_count > 0 else 0.0
    quality_pct = ((grade_counts[5] + grade_counts[4]) / total_count * 100) if total_count > 0 else 0.0
    success_pct = ((grade_counts[5] + grade_counts[4] + grade_counts[3]) / total_count * 100) if total_count > 0 else 0.0

    stat_header_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")

    stats_data = [
        ("Всего участников:", total_count),
        ("Средний процент правильных ответов:", f"{avg_pct:.1f}%"),
        ("Качественная успеваемость (оценки 4 и 5):", f"{quality_pct:.1f}%"),
        ("Абсолютная успеваемость (оценки 3, 4, 5):", f"{success_pct:.1f}%"),
        (f"Оценок «5» (от {g5}%):", f"{grade_counts[5]} чел."),
        (f"Оценок «4» (от {g4}%):", f"{grade_counts[4]} чел."),
        (f"Оценок «3» (от {g3}%):", f"{grade_counts[3]} чел."),
        (f"Оценок «2» (менее {g3}%):", f"{grade_counts[2]} чел."),
    ]

    for label, val in stats_data:
        ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=4)
        c_label = ws.cell(row=row_idx, column=2, value=label)
        c_label.font = Font(name="Calibri", size=10, bold=True, color="475569")
        c_label.alignment = Alignment(horizontal="left", vertical="center")
        c_label.fill = stat_header_fill
        c_label.border = cell_border

        ws.cell(row=row_idx, column=3).border = cell_border
        ws.cell(row=row_idx, column=3).fill = stat_header_fill
        ws.cell(row=row_idx, column=4).border = cell_border
        ws.cell(row=row_idx, column=4).fill = stat_header_fill

        c_val = ws.cell(row=row_idx, column=5, value=val)
        c_val.font = Font(name="Calibri", size=10, bold=True, color="0F172A")
        c_val.alignment = Alignment(horizontal="center", vertical="center")
        c_val.fill = stat_header_fill
        c_val.border = cell_border

        ws.row_dimensions[row_idx].height = 20
        row_idx += 1

    # --- 5. Автоподбор ширины столбцов ---
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row >= header_row_idx and cell.value is not None:
                cell_len = len(str(cell.value))
                if cell_len > max_len:
                    max_len = cell_len
        ws.column_dimensions[col_letter].width = max(max_len + 4, 10)

    ws.column_dimensions["A"].width = 6   # №
    ws.column_dimensions["B"].width = max(ws.column_dimensions["B"].width, 30)  # ФИО
    ws.column_dimensions["C"].width = max(ws.column_dimensions["C"].width, 14)  # Группа
    ws.column_dimensions["D"].width = max(ws.column_dimensions["D"].width, 24)  # Тест
    ws.column_dimensions["E"].width = max(ws.column_dimensions["E"].width, 12)  # Баллы
    ws.column_dimensions["F"].width = max(ws.column_dimensions["F"].width, 12)  # Процент
    ws.column_dimensions["G"].width = max(ws.column_dimensions["G"].width, 12)  # Оценка
    ws.column_dimensions["H"].width = max(ws.column_dimensions["H"].width, 20)  # Дата

    last_data_row = 4 + len(results)
    if len(results) > 0:
        ws.auto_filter.ref = f"A4:H{last_data_row}"

    wb.save(str(output_path))
    return str(output_path)
